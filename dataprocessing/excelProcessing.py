"""
Hacemos una clase porque se trata de varias funciones que están relacionadas y
tienen que operar cosas en comun (datafradate, paths, etc) y es mejor que esten
como atributos antes que estar pasando un monton de parámetros entre sí.
"""

# Standard library imports
import datetime
import time
import shutil
import traceback
import os

# Third party imports
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Proyect imports
from data import files_paths as fp
from data.directory import TODAY, DIRECTORY
from utilities.dfs_format import DataframesFormat
from menu.menu import (msg_archivo_faltante,
                       msg_colocar_mayor,
                       msg_mail_file_generated,
                       msg_mail_file_not_generated,
                       msg_debitados_generated,
                       msg_no_debitados,
                       msg_debitos_creditos_generated,
                       msg_no_debitos_creditos,
                       msg_cerrar_archivo,
                       msg_report_saldos_generated,
                       msg_empty_table)
from emails.mailing import send_email

DATE_FORMAT = "%d/%m/%Y"
TODAY_dt64 = datetime.datetime.today().date()
DAYS_AGO_dt64 = TODAY_dt64 - datetime.timedelta(days = 5)

# Las fechas de los df son del tipo datetime64. Acá con el .date() obtenes una variable del tipo date, y necesitas datetime64 porque no podes comparar fechas sino por ser de diferente tipo de dato.
TODAY_dt64 = np.datetime64(TODAY_dt64)
DAYS_AGO_dt64 = np.datetime64(DAYS_AGO_dt64)
CHEQUES_DEBITADOS = ['ECHEQ','Pago de Cheque', 'Cheque Camara Bs.As. 48 Hs.']
GASTOS = ['Com. mantenimiento cuenta', 'Com. regularizac administ',
        'Com.pago chq por caja', 'Com.pedido chequeras',
        'Comision Cheque Pagado por Clearing',
        'Comision por Acreditacion de Valores',
        'Comision por Depositos chqs - Remoto',
        'Comision por Transferencia B. INTERNET COM. USO-000277242',
        'Impuesto Ley 25.413 Ali Gral s/Creditos', 'Intereses por Saldo Deudor',
        'IVA - Alic Adic RG 2408 Sobre Intereses',
        'IVA - Alicuota Exceptuado Percepcion', 'IVA - Alicuota Inscripto',
        'IVA - ALICUOTA INSCRIPTO REDUCIDA',
        'IVA-Ali Adic RG 2408 s/Comis-Rec Gastos',
        'Percep Ingr Brutos Bs.As. - Categoria 1',
        'Percep Ingr Brutos Bs.As. - Condicion Q',
        'Percep Ingr Brutos Bs.As. - Condicion S',
        'Suscripcion al Periodico Accion', 'Impuesto Ley 25.413 Ali Gral s/Debitos',
        'Comision E-CHEQ pagado por Clearing', 'Servicio Modulo Pyme y Agro',
        'Recaudacion ARBA - PBA', 'Percepcion I. Brutos Categoria D Bs.As.',
        'Servicio acreditaciones automaticas SERV ACRED AUTOMATIC-',
        'Imp. Sellos - Alcanzado - Buenos Aires',
        'I.V.A. - Debito Fiscal 10,5%',
        'Reg. de Recaudacion SIRCREB-Convenio',
        'Imp.debitos tasa general']

class ExcelProcessing:
    formatter = DataframesFormat()

    def __init__(self, datos_empresa, mes, year):
        # Defino variables
        self.path_new_file = []
        self.empresa = datos_empresa["empresa"]
        self.banco = datos_empresa["banco"]
        self.path_macro = fp.paths["file_macro"]
        self.path_libro_mayor = fp.paths['file_tct']
        self.mes = mes
        self.day = datetime.datetime.today().strftime('%d')
        self.year = year
        self.paths_extractos = datos_empresa["path"]
        self.cuentas = datos_empresa["cuenta"]
        self.dfs_debitos_creditos = {}
        self.saldos = {"CREDICOOP":{}, "COMAFI":{}, "PROVINCIA":{}}

    def report_saldos(self):
        path_new_file = fp.paths["saldos"]
        bancos = list(self.saldos.keys())

        with pd.ExcelWriter(path_new_file) as writer:
            for banco in bancos:
                df = pd.DataFrame(self.saldos[banco])
                df.fillna("-", inplace = True)
                styling_df(df, banco, writer, True)

        msg_report_saldos_generated()

    def report_cheques(self):
        # while True:
            try:
                for item in range(len(self.cuentas)):
                    time.sleep(3) # Doy un tiempo para terminar la descarga del archivo, ya que puede ser que se busca tan rapido que no llego a estar en el directorio al momento de entrar a esta funcion.
                    self.path_extracto = self.paths_extractos[item]
                    self.cuenta = self.cuentas[item]
                    df = ExcelProcessing.formatter.get_bank_df(self.path_extracto)
                    # Filtro los cheques debitados
                    filter_day = (df['Fecha'].dt.day == int(self.day))
                    filter_cheques = df['Concepto'].str.contains('|'.join(CHEQUES_DEBITADOS))
                    df = df[filter_day & filter_cheques]
                    df['Fecha'] = df['Fecha'].dt.strftime(DATE_FORMAT)
                    # self.dfNotas.to_excel(writer, sheet_name="notas", index=False)
                    try:
                        if not df.empty:
                            path_new_file = f"{fp.paths['ch_debitados']} {self.empresa} {self.banco} {self.cuenta} {self.year}-{self.mes}.xlsx"
                            with pd.ExcelWriter(path_new_file) as writer:
                                styling_df(df, "Cheques debitados", writer, False)
                            msg_debitados_generated(self.empresa, self.banco, self.cuenta)
                        else:
                            msg_no_debitados(self.empresa, self.banco, self.cuenta)
                    except PermissionError:
                        msg_cerrar_archivo()
                # break
            except FileNotFoundError:
                # Se entiende que la descarga es automatica y anda, lo manda al directorio correcto. Entonces si salta este error es xq no se descargo extracto coop por no haber movmimientos. Cuando no hay movimientos no te deja descargar archivo la pagina
                msg_no_debitados(self.empresa, self.banco, self.cuenta)


    def report_debitos_creditos(self, ingreso=0):
        # while True:
            try:
                for item in range(len(self.cuentas)):
                    time.sleep(3) # Doy un tiempo para terminar la descarga del archivo, ya que puede ser que se busca tan rapido que no llego a estar en el directorio al momento de entrar a esta funcion.
                    self.path_extracto = self.paths_extractos[item]
                    if ingreso:
                        self.path_extracto = f"{DIRECTORY}\\paperworks\\movimientos_cuenta {self.empresa}.xlsx"
                    self.cuenta = self.cuentas[item]
                    df = ExcelProcessing.formatter.get_bank_df(self.path_extracto)
                    # Filtro los cheques debitados
                    df.drop("Saldo", axis = 1, inplace = True)
                    filter_day = (df['Fecha'] <= TODAY_dt64) & (df['Fecha'] >= DAYS_AGO_dt64)
                    filter_gastos = ~df['Concepto'].str.contains('|'.join(GASTOS))
                    df = df[filter_day & filter_gastos]
                    df['Fecha'] = df['Fecha'].dt.strftime(DATE_FORMAT)
                    if not df.empty:
                        if len(self.dfs_debitos_creditos) == 0 or self.empresa not in list(self.dfs_debitos_creditos.keys()):
                            self.dfs_debitos_creditos[self.empresa] = {f'{self.banco} {self.cuenta}': df}
                        elif self.empresa in list(self.dfs_debitos_creditos.keys()):
                            self.dfs_debitos_creditos[self.empresa][f'{self.banco} {self.cuenta}'] = df
                # break
            except FileNotFoundError:
                # Se entiende que la descarga es automatica y anda, lo manda al directorio correcto. Entonces si salta este error es xq no se descargo extracto coop por no haber movmimientos. Cuando no hay movimientos no te deja descargar archivo la pagina
                msg_no_debitos_creditos(self.empresa, self.banco, self.cuenta)


    def export_debitos_creditos(self):
        if len(self.dfs_debitos_creditos) > 0:
            for empresa in list(self.dfs_debitos_creditos.keys()):
                path_new_file = f"{fp.paths['debitos_creditos']} {empresa} {self.year}-{self.mes}-{self.day}.xlsx"
                self.path_new_file.append(path_new_file)
                with pd.ExcelWriter(path_new_file) as writer:
                    for cuenta in list(self.dfs_debitos_creditos[empresa].keys()):
                        styling_df(self.dfs_debitos_creditos[empresa][cuenta], cuenta, writer, False)


    def clasificados_debitos_creditos(self, EMAIL_ADDRESSES):
        for clasificacion in EMAIL_ADDRESSES.keys():
            addresses = []
            file_clasificado = False
            try:
                path_new_file = f"{DIRECTORY}\\{clasificacion} {TODAY_dt64}.xlsx"
                for path in self.path_new_file:
                    file = pd.ExcelFile(path)
                    sheets_names = file.sheet_names
                    index_mov_bancarios = path.rfind("Movimientos bancarios")
                    empresa = path[(index_mov_bancarios+22):-19]
                    for sheet in sheets_names:
                        # Hay que releer los archivos porque teoricamente se clasificaron
                        df = pd.read_excel(path, sheet_name = sheet)
                        sheet = f"{empresa} {sheet}"
                        # La columna "Clasificación" se toma con dtype="float64" por default cuando no tiene datos. Entonces si se le aplica "str.contains()" a la columna salta error- "AttributeError: Can only use .str accessor with string values!"
                        # Para evitar este error se convierte la columna a string primero
                        df["Clasificación"] = df["Clasificación"].astype(str)
                        # elimino columnas que no interesan
                        df = df.loc[:, ~df.columns.str.match("Unnamed")].copy()
                        # na = False porque sino las celdas que estan sin texto salen NaN y para filtrar se necesita una lista con valores True y False, pero no NaN.
                        filter = df["Clasificación"].str.contains(clasificacion, na=False)
                        df_aux = df[filter]
                        # Si ninguno pasa el filtro anterior no se crea el excel
                        if not df_aux.empty:
                            file_clasificado = os.path.isfile(path_new_file)
                            if not file_clasificado:
                                with pd.ExcelWriter(path_new_file, engine='openpyxl') as writer:
                                    styling_df(df_aux, sheet, writer, False)
                                file_clasificado = os.path.isfile(path_new_file)
                            if file_clasificado:
                                workbook = load_workbook(path_new_file)
                                with pd.ExcelWriter(path_new_file, engine='openpyxl') as writer:
                                    writer.book = workbook
                                    writer.sheets = {
                                        worksheet.title: worksheet
                                        for worksheet in workbook.worksheets
                                    }
                                    styling_df(df_aux, sheet, writer, False)
                            if empresa == "IKAIKA" and clasificacion == "COBRANZAS":
                                addresses.append("mendozaikaika.admi@gmail.com")

                if file_clasificado:
                    # Si se creó el Excel se envía mail
                    msg_mail_file_generated(clasificacion, "", "")
                    addresses.extend(EMAIL_ADDRESSES[clasificacion])
                    print(addresses)
                    # pathFiles, receiver, subject
                    send_email(path_new_file, addresses, f"REPORTE DIARIO - {clasificacion}")
                else:
                    msg_mail_file_not_generated(clasificacion, "", "")
            except FileNotFoundError:
                msg_archivo_faltante()

    def conciliacion_file(self, ingreso=0):
        # Usamos "while True" ya que si el archivo esta abierto, la idea es poder ejecutar de nuevo la función.
        while True:
            try:
                for item in range(len(self.cuentas)):
                    # Defino variables
                    self.cuenta = self.cuentas[item]
                    self.path_extracto = self.paths_extractos[item]
                    if ingreso:
                        self.path_extracto = f"{DIRECTORY}\\paperworks\\movimientos_cuenta {self.empresa}.xlsx"
                    path_new_file = f"{fp.paths['conciliado']}{self.empresa} {self.banco} {self.cuenta} {self.year}-{self.mes}.xlsm"
                    self.path_new_file.append(path_new_file)

                    msg_colocar_mayor(self.empresa, self.banco, self.cuenta, item)

                    if item > 0:
                        self.path_libro_mayor = self.path_libro_mayor.replace("Libro1.xlsx", f"Libro{1+item}.xlsx")
                    time.sleep(3)  # Doy un tiempo para terminar la descarga del archivo, ya que puede ser que se busca tan rapido que no llego a estar en el directorio al momento de entrar a esta funcion.

                    # Creo dataframes
                    file_extracto = os.path.isfile(self.path_extracto)
                    if file_extracto:
                        self.df_bank = ExcelProcessing.formatter.get_bank_df(self.path_extracto)
                        filter_month = (self.df_bank['Fecha'].dt.month == int(self.mes))
                        self.df_bank = self.df_bank[filter_month].copy()
                        self.df_bank['Fecha'] = self.df_bank['Fecha'].dt.strftime(DATE_FORMAT)
                    else:
                        data_empty_bank = {'Clasificación':[],
                        'Fecha':[],
                        'Concepto':[],
                        'Descripción Ampliada':[],
                        'Comprobante':[],
                        'Importe':[],
                        'Saldo':[]
                        }
                        msg_empty_table('Extracto Bancario', self.empresa, self.banco, self.cuenta)
                        self.df_bank = pd.DataFrame(data_empty_bank)

                    file_mayor = os.path.isfile(self.path_libro_mayor)
                    if file_mayor:
                        self.df_tct = pd.read_excel(self.path_libro_mayor)
                        self.df_tct['Fecha'] = self.df_tct['Fecha'].dt.strftime(DATE_FORMAT)
                        self.df_tct["Razón Social Cliente/Proveedor"].fillna("-", inplace = True)
                    else:
                        data_empty_tct = {'Número Asiento':[],
                        'Descripción':[],
                        'Fecha':[],
                        'Debe':[],
                        'Haber':[],
                        'Saldo':[],
                        'Razón Social Cliente/Proveedor':[]
                        }
                        msg_empty_table('Libro Mayor', self.empresa, self.banco, self.cuenta)
                        self.df_tct = pd.DataFrame(data_empty_tct)
                    self.df_tct.insert(0, "Clasificación", "")
                    # Copio archivo Excel Macro a directorio
                    shutil.copy(self.path_macro, self.path_new_file[item])
                    # Copio dataframe en archivo excel
                    workbook = load_workbook(self.path_new_file[item], keep_vba = True)
                    with pd.ExcelWriter(self.path_new_file[item], engine='openpyxl') as writer:
                        writer.book = workbook
                        writer.sheets = {
                            worksheet.title: worksheet
                            for worksheet in workbook.worksheets
                        }
                        self.df_bank.to_excel(writer, sheet_name="Banco", index=False, freeze_panes=(1,0))
                        self.df_tct.to_excel(writer, sheet_name="Táctica", index=False, freeze_panes=(1,0))
                break

            except PermissionError:
                print(traceback.format_exc())
                msg_cerrar_archivo()
                self.path_new_file = []


    def clasificados_conciliacion(self, EMAIL_ADDRESSES):
        for path in self.path_new_file:
            while True:
                try:
                    file = pd.ExcelFile(path)
                    sheets_names = file.sheet_names
                    for clasificacion in EMAIL_ADDRESSES.keys():
                        file_clasificado = False
                        path_new_file = path.replace(".xlsm", f" - {clasificacion}.xlsx")
                        for sheet in sheets_names:
                            # Hay que releer los archivos porque teoricamente se clasificaron
                            df = pd.read_excel(path, sheet_name = sheet)
                            # La columna "Clasificación" se toma con dtype="float64" por default cuando no tiene datos. Entonces si se le aplica "str.contains()" a la columna salta error- "AttributeError: Can only use .str accessor with string values!"
                            # Para evitar este error se convierte la columna a string primero
                            df["Clasificación"] = df["Clasificación"].astype(str)
                            # elimino columnas que no interesan
                            df = df.loc[:, ~df.columns.str.match("Unnamed")].copy()
                            # na = False porque sino las celdas que estan sin texto salen NaN y para filtrar se necesita una lista con valores True y False, pero no NaN.
                            filter = df["Clasificación"].str.contains(clasificacion, na=False)
                            df_aux = df[filter]
                            # Si ninguno pasa el filtro anterior no se crea el excel
                            if not df_aux.empty:
                                file_clasificado = os.path.isfile(path_new_file)
                                if not file_clasificado:
                                    with pd.ExcelWriter(path_new_file, engine='openpyxl') as writer:
                                        styling_df(df_aux, sheet, writer, False)
                                    file_clasificado = os.path.isfile(path_new_file)
                                if file_clasificado:
                                    workbook = load_workbook(path_new_file)
                                    with pd.ExcelWriter(path_new_file, engine='openpyxl') as writer:
                                        writer.book = workbook
                                        writer.sheets = {
                                            worksheet.title: worksheet
                                            for worksheet in workbook.worksheets
                                        }
                                        styling_df(df_aux, sheet, writer, False)
                        if file_clasificado:
                            # Si se creó el Excel se envía mail
                            # Tomo mes, empresa y banco (osea nombre de archivo)
                            file_name = path[path.rfind("\\")+1:-5]
                            msg_mail_file_generated(clasificacion, self.banco, self.cuenta)
                            addresses = EMAIL_ADDRESSES[clasificacion]
                            print(addresses)
                            # pathFiles, receiver, subject
                            send_email(path_new_file, addresses, f"REVISION BANCOS - {file_name}")
                        else:
                            msg_mail_file_not_generated(clasificacion, self.banco, self.cuenta)
                    break
                except FileNotFoundError:
                    msg_archivo_faltante()
                except PermissionError:
                    msg_cerrar_archivo()



    def calculo_saldos(self):
        # Usamos "while True" ya que si falta el archivo/esta abierto, la idea es poder ejecuar de nuevo la función.
        while True:
            try:
                for item in range(len(self.cuentas)):
                    # Defino variables
                    self.cuenta = self.cuentas[item]
                    self.path_extracto = self.paths_extractos[item]
                    path_new_file = f"{fp.paths['conciliado']}{self.empresa} {self.banco} {self.cuenta} Saldos mensuales.xlsx"
                    self.path_new_file.append(path_new_file)

                    # Creo dataframes
                    creacion_df_banco = self.switch_banco.get(self.banco)
                    creacion_df_banco("saldos")

                    ### OPERO PARA FILTRAR SALDOS POR MES QUE HAYAN EN ARCHIVO ###
                    if self.banco == "CREDICOOP":
                        # sum_debitos = self.df_bank.groupby(self.df_bank['Fecha'].dt.month)['Débito'].sum() # Agrupo solo por mes
                        # sum_creditos = self.df_bank.groupby(self.df_bank['Fecha'].dt.month)['Crédito'].sum() # Agrupo solo por mes

                        sum_debitos = self.df_bank.groupby([(self.df_bank['Fecha'].dt.year), self.df_bank['Fecha'].dt.month])['Débito'].sum() # Agrupo por mes y año
                        sum_creditos = self.df_bank.groupby([(self.df_bank['Fecha'].dt.year), self.df_bank['Fecha'].dt.month])['Crédito'].sum() # Agrupo por mes y año

                        saldos = sum_creditos - sum_debitos

                    elif self.banco == "COMAFI" or self.banco == "PROVINCIA":
                        # saldos = self.df_bank.groupby(df['Fecha'].dt.month)['Importe'].sum() # Agrupo solo por mes
                        if self.banco == "COMAFI":
                            self.df_bank['Importe'] = self.df_bank['Importe'].str.replace(',','.')
                            self.df_bank['Importe'] = self.df_bank['Importe'].astype(float)

                        saldos = self.df_bank.groupby([(self.df_bank['Fecha'].dt.year), (self.df_bank['Fecha'].dt.month)])['Importe'].sum() # Agrupo por mes y año

                    self.df_saldos = saldos.to_frame(name = "SALDOS BANCO")
                    # self.df_saldos = self.df_saldos.rename(columns = {'Fecha':'MES'}) No ANDA PARA CAMBIAR EL NOMBREEEE NO SE COMO HACER. POR AHI TIENE QUE VER QUE ES SOLKO UNA COLUMNA

                    with pd.ExcelWriter(self.path_new_file[item]) as writer:
                        styling_df(self.df_saldos, "Saldos", writer, True)

                break
            except FileNotFoundError:
                print(traceback.format_exc())
                msg_archivo_faltante()
                self.path_new_file = []
            except PermissionError:
                print(traceback.format_exc())
                msg_cerrar_archivo()
                self.path_new_file = []





def styling_df(dataframe, sheetName, writer, index):

    dataframe.to_excel(writer, sheet_name=sheetName, index=index, freeze_panes=(1,0))
    worksheet = writer.sheets[sheetName]
    #Columns AutoFit
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width
